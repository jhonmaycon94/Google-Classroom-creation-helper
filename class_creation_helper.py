import unicodedata
import re
import pyperclip
import openpyxl


class GoogleClassCreationHelper:

    @staticmethod
    def normatize(students):
        """ Recebe uma lista de nomes de estudantes
        e retorna os nomes sem acentos e cedilhas"""
        for indice, student in enumerate(students):
            student = unicodedata.normalize(
                "NFKD", student.strip()).encode(
                    "ascii", "ignore").decode("ascii")
            student = student.upper().strip()
            students[indice] = student
        return students

    def extract_students_from_SIGAA(self, clipboard_copy):
        """ Recebe a tabela de listagem dos alunos da turma no SIGAA
        e retorna somente o nome dos alunos"""
        students = re.findall(r"(\b[A-Z][^\d]{1,}\t\t)", clipboard_copy)
        students = self.normatize(students)
        if not students:
            clipboard_copy = self.normatize(str(clipboard_copy).split("\n"))
            clipboard_copy = "\n".join(clipboard_copy)
            students = re.findall(r"((?<=\d\t)[A-Z ]{3,})", clipboard_copy)
        return students

    @staticmethod
    def load_students_from_sheet():
        """Carrega estudantes da planilha
        que serve como base de dados dos emails dos alunos"""
        workbook = openpyxl.load_workbook("emails academicos.xlsx")
        sheet = workbook['Planilha1']
        students = []
        last_row = sheet.max_row
        for i in range(1, last_row):
            student = sheet.cell(row=i, column=1).value
            students.append(student)
        return students

    @staticmethod
    def search_student_in_sheet(students_from_copy, students_from_sheet):
        """ procura os estudantes extraídos da cópia do SIGAA
        na planilha com os emails"""
        students = []
        for student in students_from_copy:
            students.extend(
                [
                    student_ws
                    for student_ws in students_from_sheet
                    if student in student_ws
                ]
            )
        return students

    @staticmethod
    def students_without_email(students):
        """Verifica se o email do estudante está na planilha"""
        no_email = []
        for student in students:
            if "@" not in student:
                no_email.append(student)
        return no_email

    @staticmethod
    def students_to_clipboard(students):
        """cola a lista de estudantes para área de transferência"""
        pyperclip.copy("\n".join(students))

    def execute(self):
        students_from_sheet = self.normatize(self.load_students_from_sheet())
        students = self.extract_students_from_SIGAA(pyperclip.paste())

        self.students = self.search_student_in_sheet(
            students, students_from_sheet)
        self.students_without_email = self.students_without_email(
            self.students)
        self.students_to_clipboard(self.students)


if __name__ == "__main__":
    obj = GoogleClassCreationHelper()
    obj.execute()
    if obj.students:
        print("Alunos copiados para Área de Transferência:\n")
        for student in obj.students:
            print(student)
        if obj.students_without_email:
            print("Alunos sem email:\n")
            for student in obj.students_without_email:
                print(student)
    else:
        print("ERRO! Nenhum estudante copiado para área de trasferência")
